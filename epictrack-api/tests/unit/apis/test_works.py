# Copyright Â© 2019 Province of British Columbia
#
# Licensed under the Apache License, Version 2.0 (the 'License');
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an 'AS IS' BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Test suite for Works."""
from copy import copy
from datetime import datetime, timedelta
from enum import Enum
from http import HTTPStatus
from io import BytesIO
from urllib.parse import urljoin

from openpyxl import load_workbook

from api.services.event import EventService
from api.services.event_configuration import EventConfigurationService
from api.services.phaseservice import PhaseService
from api.models.project import Project as ProjectModel
from api.models.work_type import WorkType as WorkTypeModel
from api.models.event_template import EventTemplateVisibilityEnum
from api.services.role import RoleService
from tests.utilities.factory_scenarios import TestRoleEnum, TestWorkFirstNationEnum, TestWorkInfo, TestWorkNotesEnum
from tests.utilities.factory_utils import (
    factory_first_nation_model,
    factory_staff_model,
    factory_staff_work_role_model,
    factory_work_first_nation_model,
    factory_work_model,
)
from tests.utilities.helpers import prepare_work_payload
from api.utils import util


API_BASE_URL = "/api/v1/"


def test_get_works(client, auth_header):
    """Test get works."""
    url = urljoin(API_BASE_URL, "works")
    factory_work_model()
    factory_work_model(TestWorkInfo.work_in_active.value)
    result = client.get(url, headers=auth_header)
    in_active_works = [work for work in result.json if work['is_active'] is False]
    assert len(in_active_works) > 0
    assert result.status_code == HTTPStatus.OK


def test_get_active_works(client, auth_header):
    """Test get works."""
    url = urljoin(API_BASE_URL, "works?is_active=true")
    factory_work_model()
    factory_work_model(TestWorkInfo.work_in_active.value)
    result = client.get(url, headers=auth_header)
    in_active_works = [work for work in result.json if work['is_active'] is False]
    assert len(in_active_works) == 0
    assert result.status_code == HTTPStatus.OK


def test_create_work(client, auth_header):
    """Test create new work."""
    url = urljoin(API_BASE_URL, "works")
    payload = prepare_work_payload()
    expected_title = _extract_title(payload)
    # Scenario 1: Valid payload
    response = client.post(url, json=payload, headers=auth_header)
    response_json = response.json
    assert response.status_code == HTTPStatus.CREATED
    assert "id" in response_json
    assert expected_title == response_json.get('title')
    assert payload.get('simple_title') == response_json.get('simple_title')
    for key, value in payload.items():
        if key == "start_date":
            start_date = datetime.fromisoformat(response_json[key])
            value = datetime.fromisoformat(value)
            assert value.date() == start_date.date()
        else:
            assert value == response_json[key]
    # Scenario 2: Missing required fields


def _extract_title(payload):
    project_name = ProjectModel.find_by_id(payload.get('project_id')).name
    work_type = WorkTypeModel.find_by_id(payload.get('work_type_id')).name
    expected_title = util.generate_title(project_name, work_type, payload.get('simple_title'))
    return expected_title


def test_validate_work(client, auth_header):
    """Test validate work"""
    url = urljoin(API_BASE_URL, "works/exists")
    # Scenario 1: Updating an existing work
    work = factory_work_model()
    payload = {"title": work.title, "work_id": work.id}
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert not response.json["exists"]
    # Scenario 2: Creating new work with existing name
    del payload["work_id"]
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert response.json["exists"] is False

    # Scenario 3: Creating new work with new name
    payload = TestWorkInfo.validation_work.value
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert not response.json["exists"]


def test_get_work_details(client, auth_header):
    """Test get work"""
    work = factory_work_model()

    title = util.generate_title(work.project.name, work.work_type.name, work.simple_title)
    url = urljoin(API_BASE_URL, f"works/{work.id}")
    response = client.get(url, headers=auth_header)
    response_json = response.json
    assert title == response.json.get('title')

    assert response.status_code == HTTPStatus.OK
    for key, value in response_json.items():
        if isinstance(value, dict):
            # Ignore nested objects
            continue
        obj_value = getattr(work, key)
        if isinstance(obj_value, datetime):
            obj_value = obj_value.isoformat()
        if isinstance(obj_value, Enum):
            obj_value = obj_value.value
        assert value == obj_value


def test_work_plan(client, auth_header):
    """Test work plan"""
    payload = prepare_work_payload()

    # Create the work
    url = urljoin(API_BASE_URL, "works")
    work_response = client.post(url, json=payload, headers=auth_header)
    work_response_json = work_response.json
    assert work_response.status_code == HTTPStatus.CREATED

    # Verify team lead
    roles = RoleService.find_all()
    team_lead_role = next((x for x in roles if x.name == "Team Lead"))
    url = urljoin(API_BASE_URL, f"works/{work_response_json['id']}/staff-roles")
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    response_json = response.json
    assert any(
        (
            x["role_id"] == team_lead_role.id and x["staff_id"] == payload["work_lead_id"]
            for x in response_json
        )
    )

    # Verify phases
    phases = PhaseService.find_phase_codes_by_ea_act_and_work_type(
        payload["ea_act_id"], payload["work_type_id"]
    )
    phase_ids = [x.id for x in phases]
    url = urljoin(API_BASE_URL, f"works/{work_response_json['id']}/phases")
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    work_phase_json = response.json
    work_phase_ids = [x["work_phase"]["phase_id"] for x in work_phase_json]
    assert all([phase_id in work_phase_ids for phase_id in phase_ids])

    for index, work_phase in enumerate(work_phase_json):
        # Verify mandatory milestones
        mandatory_event_configurations = EventConfigurationService.find_configurations(
            work_phase_id=work_phase["work_phase"]["id"], visibility_modes=[EventTemplateVisibilityEnum.MANDATORY.value]
        )
        work_milestone_events = EventService.find_milestone_events_by_work_phase(
            work_phase["work_phase"]["id"]
        )
        work_milestone_event_configurations = [
            x.event_configuration_id for x in work_milestone_events
        ]
        assert all(
            [
                mandatory_event_configuration.id in work_milestone_event_configurations
                for mandatory_event_configuration in mandatory_event_configurations
            ]
        )

        # verify start date
        if index == 0:
            assert (
                work_response_json["start_date"]
                == work_phase["work_phase"]["start_date"]
            )
        else:
            prev_work_phase = work_phase_json[index - 1]
            prev_start_date = prev_work_phase["work_phase"]["start_date"]
            prev_start_date = datetime.fromisoformat(prev_start_date)
            assert (
                work_phase["work_phase"]["start_date"] == (
                    prev_start_date + timedelta(days=prev_work_phase["total_number_of_days"] + 1)
                ).isoformat()
            )


def test_update_work(client, auth_header):
    """Test update work"""
    work = factory_work_model()
    updated_data = copy(TestWorkInfo.work1.value)
    updated_data["project_id"] = work.project_id
    updated_data["simple_title"] = "Work title updated"
    url = urljoin(API_BASE_URL, f"works/{work.id}")
    response = client.put(url, headers=auth_header, json=updated_data)
    response_json = response.json
    assert response.status_code == HTTPStatus.OK
    assert response_json["id"] == work.id
    assert response_json["simple_title"] == updated_data["simple_title"]


def test_delete_work(client, auth_header):
    """Test delete work"""
    work = factory_work_model()
    url = urljoin(API_BASE_URL, f"works/{work.id}")
    response = client.delete(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert response.text == "Work successfully deleted"

    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.NOT_FOUND


def test_work_dashboard(client, auth_header):
    """Test work dashboard"""
    payload = prepare_work_payload()

    # Create the work
    url = urljoin(API_BASE_URL, "works")
    work_response = client.post(url, json=payload, headers=auth_header)
    work_response_json = work_response.json
    assert work_response.status_code == HTTPStatus.CREATED

    url = urljoin(API_BASE_URL, "works/dashboard")
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert len(response.json["items"]) == 1
    assert response.json["total"] == 1
    assert response.json["items"][0]["id"] == work_response_json["id"]


def test_work_resources(client, auth_header):
    """Test work resources"""
    payload = prepare_work_payload()

    # Create the work
    url = urljoin(API_BASE_URL, "works")
    work_response = client.post(url, json=payload, headers=auth_header)
    work_response_json = work_response.json
    assert work_response.status_code == HTTPStatus.CREATED

    url = urljoin(API_BASE_URL, "works/resources?is_active=true")
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    work_resource_json = response.json[0]
    assert work_resource_json["eao_team"]["id"] == work_response_json["eao_team_id"]
    assert work_resource_json["project"]["id"] == work_response_json["project_id"]
    assert (
        work_resource_json["responsible_epd"]["id"]
        == work_response_json["responsible_epd_id"]
    )
    assert work_resource_json["work_lead"]["id"] == work_response_json["work_lead_id"]
    assert len(work_resource_json["staff"]) == 2
    team_lead = [x for x in work_resource_json["staff"] if x["role"]['name'] == 'Team Lead'][0]
    epd = [x for x in work_resource_json["staff"] if x["role"]['name'] == 'Responsible EPD'][0]
    assert team_lead["id"] == work_response_json["work_lead_id"]
    assert epd["id"] == work_response_json["responsible_epd_id"]


def test_get_work_staff_roles(client, auth_header):
    """Test get work staff roles"""
    payload = prepare_work_payload()

    # Create the work
    url = urljoin(API_BASE_URL, "works")
    work_response = client.post(url, json=payload, headers=auth_header)
    work_response_json = work_response.json
    assert work_response.status_code == HTTPStatus.CREATED

    url = urljoin(API_BASE_URL, f'works/{work_response_json["id"]}/staff-roles')
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert len(response.json) == 2


def test_create_work_staff_roles(client, auth_header):
    """Test create work staff roles"""
    work = factory_work_model()
    staff = factory_staff_model()
    payload = {
        "staff_id": staff.id,
        "role_id": TestRoleEnum.OFFICER_ANALYST.value,
        "is_active": True,
    }
    url = urljoin(API_BASE_URL, f"works/{work.id}/staff-roles")
    response = client.post(url, json=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.CREATED
    response_json = response.json
    assert work.id == response_json["work_id"]
    assert payload["role_id"] == response_json["role_id"]
    assert payload["staff_id"] == response_json["staff_id"]


def test_get_work_staff_role_details(client, auth_header):
    """Test get work staff role details"""
    staff_work_role = factory_staff_work_role_model()
    url = urljoin(API_BASE_URL, f"works/staff-roles/{staff_work_role.id}")
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    response_json = response.json
    assert staff_work_role.id == response_json["id"]
    assert staff_work_role.work_id == response_json["work_id"]
    assert staff_work_role.staff_id == response_json["staff_id"]
    assert staff_work_role.role_id == response_json["role_id"]


def test_update_work_staff_role_details(client, auth_header):
    """Test update work staff role details"""
    staff_work_role = factory_staff_work_role_model()
    payload = {
        "staff_id": staff_work_role.staff_id,
        "is_active": not staff_work_role.is_active,
        "role_id": TestRoleEnum.FN_CAIRT.value,
    }
    url = urljoin(API_BASE_URL, f"works/staff-roles/{staff_work_role.id}")
    response = client.put(url, json=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    response_json = response.json
    assert staff_work_role.id == response_json["id"]
    assert staff_work_role.work_id == response_json["work_id"]
    assert payload["staff_id"] == response_json["staff_id"]
    assert payload["role_id"] == response_json["role_id"]
    assert payload["is_active"] == response_json["is_active"]


def test_validate_work_staff(client, auth_header):
    """Test validate work staff"""
    work = factory_work_model()
    url = urljoin(API_BASE_URL, f"works/{work.id}/staff-roles/exists")

    # Scenario 1: Updating an existing work staff
    staff_work_role = factory_staff_work_role_model(work_id=work.id)
    payload = {
        "staff_id": staff_work_role.staff_id,
        "role_id": staff_work_role.role_id,
        "work_staff_id": staff_work_role.id,
    }
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert not response.json["exists"]

    # Scenario 2: Creating new work staff with existing data
    del payload["work_staff_id"]
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert response.json["exists"]

    # Scenario 3: Creating new work staff with new staff
    staff = factory_staff_model()
    payload["staff_id"] = staff.id
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert not response.json["exists"]


def test_download_work_plan(client, auth_header):
    """Test download work plan"""
    payload = prepare_work_payload()
    # Create the work
    url = urljoin(API_BASE_URL, "works")
    work_response = client.post(url, json=payload, headers=auth_header)
    work_response_json = work_response.json
    assert work_response.status_code == HTTPStatus.CREATED
    url = urljoin(API_BASE_URL, "works/workplan/download")
    params = {"work_phase_id": work_response_json["current_work_phase_id"]}
    headers = copy(auth_header)
    headers.update({"Response-Type": "blob"})
    response = client.post(url, query_string=params, headers=headers)
    assert response.status_code == HTTPStatus.OK
    assert (
        response.headers["Content-Disposition"] == "attachment; filename=work_plan.xlsx"
    )
    # Try and open the response to make sure it is a valid excel file
    # will raise 'zipfile.BadZipFile: File is not a zip file' exception if opening fails.
    load_workbook(filename=BytesIO(response.data))


def test_work_phase_template_upload_status(client, auth_header):
    """Test work phase template upload status"""
    payload = prepare_work_payload()
    # Create the work
    url = urljoin(API_BASE_URL, "works")
    work_response = client.post(url, json=payload, headers=auth_header)
    work_response_json = work_response.json
    assert work_response.status_code == HTTPStatus.CREATED
    url = urljoin(
        API_BASE_URL, f"works/work-phases/{work_response_json['current_work_phase_id']}/template-upload-status"
    )
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert not response.json["template_available"]
    assert not response.json["task_added"]


def test_get_work_first_nations(client, auth_header):
    """Test get work first nations"""
    payload = prepare_work_payload()

    # Create the work
    url = urljoin(API_BASE_URL, "works")
    work_response = client.post(url, json=payload, headers=auth_header)
    work_response_json = work_response.json
    assert work_response.status_code == HTTPStatus.CREATED

    url = urljoin(API_BASE_URL, f'works/{work_response_json["id"]}/first-nations')
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert len(response.json) == 0


def test_create_work_first_nation(client, auth_header):
    """Test create work first nation"""
    work = factory_work_model()
    payload = TestWorkFirstNationEnum.work_first_nation1.value
    payload["work_id"] = work.id
    payload["indigenous_nation_id"] = factory_first_nation_model().id
    url = urljoin(API_BASE_URL, f"works/{work.id}/first-nations")
    response = client.post(url, json=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.CREATED
    response_json = response.json
    for key, value in payload.items():
        assert value == response_json[key]


def test_get_work_first_nation_details(client, auth_header):
    """Test get work first nation details"""
    work_first_nation = factory_work_first_nation_model()
    url = urljoin(API_BASE_URL, f"works/first-nations/{work_first_nation.id}")
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    response_json = response.json
    assert work_first_nation.id == response_json["id"]
    assert work_first_nation.work_id == response_json["work_id"]
    assert (
        work_first_nation.indigenous_nation_id == response_json["indigenous_nation_id"]
    )
    assert (
        work_first_nation.indigenous_category_id
        == response_json["indigenous_category_id"]
    )
    assert (
        work_first_nation.indigenous_consultation_level_id
        == response_json["indigenous_consultation_level_id"]
    )
    assert work_first_nation.is_active == response_json["is_active"]


def test_update_work_first_nation_details(client, auth_header):
    """Test update work first nation details"""
    work_first_nation = factory_work_first_nation_model()
    work_first_nation.as_dict
    payload = {
        "work_id": work_first_nation.work_id,
        "indigenous_nation_id": work_first_nation.indigenous_nation_id,
        "indigenous_category_id": work_first_nation.indigenous_category_id,
        "indigenous_consultation_level_id": work_first_nation.indigenous_consultation_level_id,
        "is_active": not work_first_nation.is_active,
    }
    url = urljoin(API_BASE_URL, f"works/first-nations/{work_first_nation.id}")
    response = client.put(url, json=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    response_json = response.json
    assert work_first_nation.id == response_json["id"]
    assert payload["is_active"] == response_json["is_active"]


def test_validate_work_first_nation(client, auth_header):
    """Test validate work nation"""
    # Scenario 1: Updating an existing work first nation
    work_first_nation = factory_work_first_nation_model()
    url = urljoin(
        API_BASE_URL, f"works/{work_first_nation.work_id}/first-nations/exists"
    )
    payload = {
        "indigenous_nation_id": work_first_nation.indigenous_nation_id,
        "work_indigenous_nation_id": work_first_nation.id,
    }
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert not response.json["exists"]

    # Scenario 2: Creating new work first nation with existing data
    del payload["work_indigenous_nation_id"]
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert response.json["exists"]

    # Scenario 3: Creating new work first nation with new first nation
    first_nation = factory_first_nation_model()
    payload["indigenous_nation_id"] = first_nation.id
    response = client.get(url, query_string=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert not response.json["exists"]


def test_download_work_first_nations(client, auth_header):
    """Test download work first nations"""
    work_first_nation = factory_work_first_nation_model()
    url = urljoin(
        API_BASE_URL, f"works/{work_first_nation.work_id}/first-nations/download"
    )
    headers = copy(auth_header)
    headers.update({"Response-Type": "blob"})
    response = client.post(url, headers=headers)
    assert response.status_code == HTTPStatus.OK
    assert (
        response.headers["Content-Disposition"]
        == "attachment; filename=first_nations.xlsx"
    )
    # Try and open the response to make sure it is a valid excel file
    # will raise 'zipfile.BadZipFile: File is not a zip file' exception if opening fails.
    load_workbook(filename=BytesIO(response.data))


def test_work_notes(client, auth_header):
    """Test save work notes"""
    work = factory_work_model()
    url = urljoin(API_BASE_URL, f"works/{work.id}/notes")

    # Scenario 1: status notes
    payload = TestWorkNotesEnum.work_notes1.value
    response = client.patch(url, json=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert response.json["status_notes"] == payload["notes"]

    # Scenario 2: issue notes
    payload = TestWorkNotesEnum.work_notes2.value
    response = client.patch(url, json=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert response.json["issue_notes"] == payload["notes"]


def test_work_first_notes(client, auth_header):
    """Test save work first notes"""
    work = factory_work_model()
    url = urljoin(API_BASE_URL, f"works/{work.id}/first-nation-notes")

    # Scenario 1: status notes
    payload = copy(TestWorkNotesEnum.work_notes1.value)
    del payload["note_type"]

    response = client.patch(url, json=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert response.json["first_nation_notes"] == payload["notes"]


def test_import_first_nations(client, auth_header):
    """Test save work first notes"""
    work = factory_work_model()
    url = urljoin(API_BASE_URL, f"works/{work.id}/first-nations/import")
    work_first_nation1 = factory_work_first_nation_model(work_id=work.id)
    work_first_nation2 = factory_work_first_nation_model(work_id=work.id)
    work_first_nation3 = factory_work_first_nation_model(work_id=work.id)
    payload = {
        "indigenous_nation_ids": [
            work_first_nation1.indigenous_nation_id,
            work_first_nation2.indigenous_nation_id,
            work_first_nation3.indigenous_nation_id,
        ]
    }

    response = client.post(url, json=payload, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
    assert response.text == "Imported successfully"


def test_get_work_types(client, auth_header):
    """Test get work types"""
    url = urljoin(API_BASE_URL, 'works/types')
    response = client.get(url, headers=auth_header)
    assert response.status_code == HTTPStatus.OK
